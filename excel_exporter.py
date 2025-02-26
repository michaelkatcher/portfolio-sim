#!/usr/bin/env python3
"""
Excel Exporter for MCA Portfolio Simulator

This module handles the exporting of portfolio simulation results to Excel,
creating a formatted workbook with multiple sheets as specified in the project requirements.
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator

from common_utils import setup_logging, ExportError
from constants import DEAL_COLUMNS, PAYMENT_COLUMNS, EXCEL

# Set up logger
logger = setup_logging(__name__)


def export_portfolio_to_excel(portfolio, filename, verbose=False):
    """
    Export a portfolio simulation to Excel with formatted sheets.
    
    Args:
        portfolio: A simulated Portfolio instance
        filename: The output Excel file path
        verbose: Whether to print progress information
        
    Returns:
        Path to the created Excel file
        
    Raises:
        ExportError: If export fails
        ValueError: If inputs are invalid
    """
    # Validate inputs
    if portfolio is None:
        raise ValueError("Portfolio cannot be None")
        
    if not filename:
        raise ValueError("Filename cannot be empty")
    
    # Make sure output directory exists
    try:
        output_dir = os.path.dirname(filename)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            if verbose:
                logger.info(f"Created output directory: {output_dir}")
    except Exception as e:
        raise ExportError(f"Failed to create output directory: {str(e)}") from e
    
    if verbose:
        logger.info(f"Exporting portfolio '{portfolio.name}' to Excel: {filename}")
    
    try:
        # Create a workbook
        wb = Workbook()
        
        # Remove the default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Create each sheet - order matters for formulas
        _create_deals_sheet(wb, portfolio)
        _create_cashflows_sheet(wb, portfolio)
        # Portfolio scenario sheet must be created last since it references other sheets
        _create_portfolio_scenario_sheet(wb, portfolio)
        
        # Save the workbook
        wb.save(filename)
        
        if verbose:
            logger.info(f"Export completed: {filename}")
        
        return filename
        
    except Exception as e:
        raise ExportError(f"Failed to export portfolio to Excel: {str(e)}") from e


# =================== Helper Functions ===================

def _set_column_widths(ws, width_map):
    """
    Set column widths from a map of column letters to widths.
    
    Args:
        ws: Worksheet object
        width_map: Dict mapping column letters to widths
    """
    try:
        for col, width in width_map.items():
            ws.column_dimensions[col].width = width
    except Exception as e:
        logger.warning(f"Error setting column widths: {str(e)}")


def _apply_header_style(cell, is_section_header=False):
    """
    Apply standard header styling to a cell.
    
    Args:
        cell: Cell object to style
        is_section_header: Whether this is a section header (different style)
    
    Returns:
        The styled cell
    """
    try:
        if is_section_header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9D9D9")
        else:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        
        cell.alignment = Alignment(horizontal='center')
        return cell
    except Exception as e:
        logger.warning(f"Error applying header style: {str(e)}")
        return cell


def _format_money_cell(cell, value):
    """
    Format a cell as currency.
    
    Args:
        cell: Cell object to format
        value: Value to set
    
    Returns:
        The formatted cell
    """
    try:
        cell.value = value
        cell.number_format = '"$"#,##0_);("$"#,##0);"-"'
        return cell
    except Exception as e:
        logger.warning(f"Error formatting money cell: {str(e)}")
        cell.value = value  # At least set the value even if formatting fails
        return cell


def _format_percent_cell(cell, value):
    """
    Format a cell as percentage.
    
    Args:
        cell: Cell object to format
        value: Value to set (0-1 range)
    
    Returns:
        The formatted cell
    """
    try:
        cell.value = value
        cell.number_format = '0.00%'
        return cell
    except Exception as e:
        logger.warning(f"Error formatting percent cell: {str(e)}")
        cell.value = value  # At least set the value even if formatting fails
        return cell


def _format_date_cell(cell, value):
    """
    Format a cell as date.
    
    Args:
        cell: Cell object to format
        value: Value to set (datetime object or string)
    
    Returns:
        The formatted cell
    """
    try:
        # Important: Pass the datetime object directly to Excel 
        # rather than converting to string first
        cell.value = value
        # Use Excel's built-in date format
        cell.number_format = 'yyyy-mm-dd'
        return cell
    except Exception as e:
        logger.warning(f"Error formatting date cell: {str(e)}")
        cell.value = value  # At least set the value even if formatting fails
        return cell


def _format_boolean_cell(cell, value):
    """
    Format a cell as a Yes/No boolean value.
    
    Args:
        cell: Cell object to format
        value: Boolean value
    
    Returns:
        The formatted cell
    """
    try:
        cell.value = "Yes" if value else "No"
        cell.alignment = Alignment(horizontal='center')
        return cell
    except Exception as e:
        logger.warning(f"Error formatting boolean cell: {str(e)}")
        cell.value = "Yes" if value else "No"  # At least set the value even if formatting fails
        return cell


def _add_section_header(ws, row, text, span=2):
    """
    Add a section header to a worksheet.
    
    Args:
        ws: Worksheet object
        row: Row number
        text: Header text
        span: Number of columns to span
    
    Returns:
        Next available row number
    """
    try:
        ws.cell(row=row, column=1, value=text)
        _apply_header_style(ws.cell(row=row, column=1), True)
        
        # Apply styling to additional columns if span > 1
        for col in range(2, span + 1):
            _apply_header_style(ws.cell(row=row, column=col), True)
        
        # Merge cells if span > 1
        if span > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        
        return row + 1
    except Exception as e:
        logger.warning(f"Error adding section header: {str(e)}")
        return row + 1  # Return next row anyway to continue


def _add_label_value_row(ws, row, label, value=None, value_format=None, formula=None):
    """
    Add a label-value pair row to a worksheet.
    
    Args:
        ws: Worksheet object
        row: Row number
        label: Label for column A
        value: Value for column B (None if using formula)
        value_format: Optional function to format the value cell
        formula: Optional Excel formula string instead of value
    
    Returns:
        Next available row number
    """
    try:
        ws.cell(row=row, column=1, value=label)
        
        value_cell = ws.cell(row=row, column=2)
        
        # If formula is provided, use it instead of value
        if formula:
            value_cell.value = formula
        elif value_format:
            value_format(value_cell, value)
        else:
            value_cell.value = value
        
        return row + 1
    except Exception as e:
        logger.warning(f"Error adding label-value row: {str(e)}")
        return row + 1  # Return next row anyway to continue


def _get_monthly_vintages(portfolio):
    """
    Extract unique monthly vintages from a portfolio.
    
    Args:
        portfolio: Portfolio instance with deal_allocations
        
    Returns:
        List of vintage months sorted chronologically
    """
    try:
        if portfolio.deal_allocations is None or DEAL_COLUMNS['FUNDING_DATE'] not in portfolio.deal_allocations.columns:
            return []
            
        # Extract month from funding date
        monthly_vintages = portfolio.deal_allocations[DEAL_COLUMNS['FUNDING_DATE']].dt.strftime('%Y-%m').unique()
        
        # Sort chronologically
        return sorted(monthly_vintages)
    except Exception as e:
        logger.warning(f"Error extracting monthly vintages: {str(e)}")
        return []


def _create_table_from_df(ws, df, start_row, columns=None, formats=None, include_totals=False):
    """
    Create a table from a DataFrame.
    
    Args:
        ws: Worksheet object
        df: DataFrame to display
        start_row: Starting row
        columns: List of columns to include (defaults to all)
        formats: Dict mapping column names to formatting functions
        include_totals: Whether to include totals row
    
    Returns:
        Next available row number
    """
    try:
        # Select columns if specified
        if columns:
            # Only include columns that exist in the DataFrame
            available_columns = [col for col in columns if col in df.columns]
            df = df[available_columns]
        else:
            available_columns = df.columns.tolist()
        
        # Set up column formats
        formats = formats or {}
        
        # Add headers
        current_row = start_row
        for col_idx, col_name in enumerate(available_columns, 1):
            # Use more readable header names
            display_name = col_name.replace('_', ' ')
            cell = ws.cell(row=current_row, column=col_idx, value=display_name)
            _apply_header_style(cell)
        
        current_row += 1
        
        # Add data rows
        for _, row_data in df.iterrows():
            for col_idx, col_name in enumerate(available_columns, 1):
                value = row_data[col_name]
                cell = ws.cell(row=current_row, column=col_idx)
                
                # Apply format if specified
                if col_name in formats:
                    formats[col_name](cell, value)
                else:
                    cell.value = value
            
            current_row += 1
        
        # Add totals row if requested
        if include_totals and not df.empty:
            ws.cell(row=current_row, column=1, value="TOTAL").font = Font(bold=True)
            
            for col_idx, col_name in enumerate(available_columns, 1):
                # Only sum numeric columns
                if col_name in formats and (formats[col_name] == _format_money_cell or 
                                           pd.api.types.is_numeric_dtype(df[col_name])):
                    total = df[col_name].sum()
                    cell = ws.cell(row=current_row, column=col_idx, value=total)
                    cell.font = Font(bold=True)
                    
                    # Use the same format as the column
                    if col_name in formats:
                        formats[col_name](cell, total)
                
                # Calculate weighted average for percentage columns
                elif col_name == DEAL_COLUMNS['COMMISSION'] and DEAL_COLUMNS['BALANCE'] in df.columns:
                    weighted_avg = (df[DEAL_COLUMNS['COMMISSION']] * df[DEAL_COLUMNS['BALANCE']]).sum() / df[DEAL_COLUMNS['BALANCE']].sum()
                    cell = ws.cell(row=current_row, column=col_idx)
                    _format_percent_cell(cell, weighted_avg)
                    cell.font = Font(bold=True)
            
            current_row += 1
        
        return current_row
    
    except Exception as e:
        logger.error(f"Error creating table from DataFrame: {str(e)}")
        return start_row + 1  # Return a row number to continue


# =================== Sheet Creation Functions ===================

def _create_portfolio_scenario_sheet(wb, portfolio):
    """
    Create the Portfolio Scenario worksheet.
    
    Args:
        wb: Workbook object
        portfolio: Portfolio instance
        
    Returns:
        Worksheet object
    """
    try:
        ws = wb.create_sheet("Portfolio Scenario", 0)
        
        # Set column widths
        _set_column_widths(ws, {'A': 30, 'B': 50})
        
        # Add title
        ws.cell(row=1, column=1, value="Portfolio Scenario")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Current row tracker
        row = 3
        
        # Add selection criteria section
        row = _add_section_header(ws, row, "Selection Criteria")
        
        # Product Types
        product_types = ", ".join(portfolio.selection_criteria.get('product_types', [])) if portfolio.selection_criteria.get('product_types') else "All"
        row = _add_label_value_row(ws, row, "Product Types", product_types)
        
        # Credit Grades
        credit_tiers = ", ".join(portfolio.selection_criteria.get('credit_tiers', [])) if portfolio.selection_criteria.get('credit_tiers') else "All"
        row = _add_label_value_row(ws, row, "Credit Grades", credit_tiers)
        
        # Deal Size Range
        if portfolio.selection_criteria.get('deal_size_range'):
            min_size, max_size = portfolio.selection_criteria['deal_size_range']
            deal_size_range = f"${min_size:,.2f} to ${max_size:,.2f}"
        else:
            deal_size_range = "All"
        row = _add_label_value_row(ws, row, "Deal Size Range", deal_size_range)
        
        # Vintages
        if portfolio.selection_criteria.get('vintage_range'):
            start_date, end_date = portfolio.selection_criteria['vintage_range']
            if isinstance(start_date, datetime):
                start_date = start_date.strftime('%Y-%m-%d')
            if isinstance(end_date, datetime):
                end_date = end_date.strftime('%Y-%m-%d')
            vintage_range = f"{start_date} to {end_date}"
        else:
            vintage_range = "All"
        row = _add_label_value_row(ws, row, "Vintages", vintage_range)
        
        # Add allocation policy section
        row += 1
        row = _add_section_header(ws, row, "Allocation Policy")
        
        # Allocation Formula
        min_amt = portfolio.allocation_policy.min_amount or 0
        max_amt = portfolio.allocation_policy.max_amount or "∞"
        pct = portfolio.allocation_policy.percentage * 100
        
        if min_amt > 0 and max_amt != "∞":
            formula_str = f"MAX({min_amt:,.0f}, MIN({max_amt:,.0f}, {pct:.1f}%*DealSize))"
        elif min_amt > 0:
            formula_str = f"MAX({min_amt:,.0f}, {pct:.1f}%*DealSize)"
        elif max_amt != "∞":
            formula_str = f"MIN({max_amt:,.0f}, {pct:.1f}%*DealSize)"
        else:
            formula_str = f"{pct:.1f}%*DealSize"
        
        row = _add_label_value_row(ws, row, "Allocation Formula", formula_str)
        
        # Add fee structure section
        row += 1
        row = _add_section_header(ws, row, "Fee Structure")
        
        # Management Fee
        fee_pct = portfolio.fee_schedule.percentage * 100
        if portfolio.fee_schedule.apply_after_principal:
            fee_desc = f"{fee_pct:.1f}% of incoming cashflows after principal repayment"
        else:
            fee_desc = f"{fee_pct:.1f}% of all incoming cashflows"
        row = _add_label_value_row(ws, row, "Management Fee", fee_desc)
        
        # Add portfolio basic information
        if portfolio.deal_allocations is not None:
            row += 1
            row = _add_section_header(ws, row, "Portfolio Summary")
            
            # Total Deals - using formula that counts rows in Deals sheet (excluding header)
            deals_formula = "=COUNTA(Deals!A:A)-1"
            row = _add_label_value_row(ws, row, "Total Deals", formula=deals_formula)
            
            # Find columns for use in formulas
            cashflows_sheet = wb["Cashflows"]
            deals_sheet = wb["Deals"]
            
            # Find allocation amount column in Deals sheet
            
            for col in range(1, deals_sheet.max_column + 1):
                if deals_sheet.cell(row=1, column=col).value == "Allocation Amount":
                    allocation_col = get_column_letter(col)
                    break
                    
            # Find columns in Deal sheet
            allocation_col = None
            deal_size_col = None
            for col in range(1, deals_sheet.max_column + 1):
                header = deals_sheet.cell(row=1, column=col).value
                if header == "Allocation Amount":
                    allocation_col = get_column_letter(col)
                elif header == "Total Original Balance":
                    deal_size_col = get_column_letter(col)
                
            # Find columns in Cashflows sheet
            portfolio_amt_col = None
            date_col = None
            funded_dt_col = None
            for col in range(1, cashflows_sheet.max_column + 1):
                header = cashflows_sheet.cell(row=1, column=col).value
                if header == "Portfolio Amount":
                    portfolio_amt_col = get_column_letter(col)
                elif header == "Transaction Date":
                    date_col = get_column_letter(col)
                elif header == "Funded Date":
                    funded_dt_col = get_column_letter(col)
                    
            # Total Deal Size formula       
            # SUM formula using the found column
            deal_size_formula = f"=SUM(Deals!{deal_size_col}:{deal_size_col})"
            row = _add_label_value_row(
                ws, row, "Total Deal Size", 
                formula=deal_size_formula
            )
            # Apply currency formatting
            ws.cell(row=row-1, column=2).number_format = EXCEL['CURRENCY_FORMAT']
            
            # Total Allocated Amount formula
            if allocation_col:
                # SUM formula using the found column
                allocation_formula = f"=SUM(Deals!{allocation_col}:{allocation_col})"
                row = _add_label_value_row(
                    ws, row, "Total Allocated Amount", 
                    formula=allocation_formula
                )
                # Apply currency formatting
                ws.cell(row=row-1, column=2).number_format = EXCEL['CURRENCY_FORMAT']
            else:
                # Fallback to hardcoded value if column not found
                total_allocation = portfolio.deal_allocations['Allocation Amount'].sum()
                row = _add_label_value_row(
                    ws, row, "Total Allocated Amount", total_allocation, 
                    value_format=_format_money_cell
                )
            
            # Average Deal Size formula
            avg_deal_size_formula = f"=B{row-2}/B{row-3}"
            row = _add_label_value_row(
                ws, row, "Avg Deal Size", 
                formula=avg_deal_size_formula)
            ws.cell(row=row-1, column=2).number_format = EXCEL['CURRENCY_FORMAT']
                
            # Average Allocation Amount formula
            avg_allocation_formula = f"=B{row-2}/B{row-4}"
            row = _add_label_value_row(
                ws, row, "Avg Allocation Amount", 
                formula=avg_allocation_formula)
            ws.cell(row=row-1, column=2).number_format = EXCEL['CURRENCY_FORMAT']
                
            # Average Deal % formula
            avg_deal_pct_formula = f"=B{row-1}/B{row-2}"
            row = _add_label_value_row(
                ws, row, "Avg Deal %", 
                formula=avg_deal_pct_formula)
            ws.cell(row=row-1, column=2).number_format = "0.0%"
            
            # Add IRR formula if we have cashflow data with dates
            if portfolio_amt_col and date_col:
                # IRR formula using XIRR
                irr_formula = f"=XIRR(OFFSET(Cashflows!{portfolio_amt_col}$1,1,0,COUNTA(Cashflows!{portfolio_amt_col}:{portfolio_amt_col})-1),OFFSET(Cashflows!{date_col}$1,1,0,COUNTA(Cashflows!{portfolio_amt_col}:{portfolio_amt_col})-1))"
                row = _add_label_value_row(ws, row, "IRR", formula=irr_formula)
                # Format as percentage
                ws.cell(row=row-1, column=2).number_format = "0.00%"
                
                # MOIC formula 
                moic_formula = f"=SUMIFS(Cashflows!{portfolio_amt_col}:{portfolio_amt_col},Cashflows!{portfolio_amt_col}:{portfolio_amt_col},\">0\")/-SUMIFS(Cashflows!{portfolio_amt_col}:{portfolio_amt_col},Cashflows!{portfolio_amt_col}:{portfolio_amt_col},\"<0\")"
                row = _add_label_value_row(ws, row, "MOIC", formula=moic_formula)
                # Format as number with 2 decimal places
                ws.cell(row=row-1, column=2).number_format = "0.00x"
            
            # Add Monthly MOICs section if we have vintage data
            monthly_vintages = _get_monthly_vintages(portfolio)
            if monthly_vintages and portfolio_amt_col and DEAL_COLUMNS['FUNDING_DATE'] in portfolio.deal_allocations.columns:
                row += 1
                row = _add_section_header(ws, row, "Monthly MOICs")
                
                # Add a formula for each monthly vintage
                for vintage in monthly_vintages:
                    # The formula identifies deals from the specific vintage
                    # and calculates the MOIC as (positive cashflows) / abs(negative cashflows)
                    
                    # For Excel: Get the first day of the vintage month as string
                    year, month = vintage.split('-')
                    first_day = f"{year}-{month}-01"
                    last_day = f"{year}-{month}-31"  # Using 31 to cover all possible month lengths
                    
                    # Simplified formula for Excel 2019 which doesn't support FILTER
                    vintage_formula = (
                        f"=SUMIFS(Cashflows!{portfolio_amt_col}:{portfolio_amt_col},"
                        f"Cashflows!{funded_dt_col}:{funded_dt_col},"
                            f"\">=\"&DATE(LEFT(A{row},4),RIGHT(A{row},2),1),"
                        f"Cashflows!{funded_dt_col}:{funded_dt_col},"
                            f"\"<=\"&DATE(LEFT(A{row},4),RIGHT(A{row},2)+1,1)-1,"
                        f"Cashflows!{portfolio_amt_col}:{portfolio_amt_col},\">0\")/"
                        f"ABS(SUMIFS(Cashflows!{portfolio_amt_col}:{portfolio_amt_col},"
                        f"Cashflows!{funded_dt_col}:{funded_dt_col},"
                            f"\">=\"&DATE(LEFT(A{row},4),RIGHT(A{row},2),1),"
                        f"Cashflows!{funded_dt_col}:{funded_dt_col},"
                            f"\"<=\"&DATE(LEFT(A{row},4),RIGHT(A{row},2)+1,1)-1,"
                        f"Cashflows!{portfolio_amt_col}:{portfolio_amt_col},\"<0\"))"
                    )
                    
                    # Add the formula with the vintage label
                    row = _add_label_value_row(ws, row, f"{vintage}", formula=vintage_formula)
                    ws.cell(row=row-1, column=2).number_format = "0.00x"
        
            # Add Portfolio History table to the right of the main content
            # This should be added near the end of _create_portfolio_scenario_sheet function
            # Before the final "return ws" statement

            # Set column widths for Portfolio History columns
            ws.column_dimensions['D'].width = 15  # Date
            ws.column_dimensions['E'].width = 25  # Funded
            ws.column_dimensions['F'].width = 25  # Balance

            # Add Portfolio History header (merged across columns D-F)
            history_header_row = 3
            ws.cell(row=history_header_row, column=4, value="Portfolio History")
            ws.cell(row=history_header_row, column=4).font = Font(bold=True)
            ws.cell(row=history_header_row, column=4).fill = PatternFill("solid", fgColor="D9D9D9")
            ws.merge_cells(start_row=history_header_row, start_column=4, end_row=history_header_row, end_column=6)
            ws.cell(row=history_header_row, column=4).alignment = Alignment(horizontal='center')

            # Add column headers
            header_row = history_header_row + 1
            ws.cell(row=header_row, column=4, value="Date").font = Font(bold=True)
            ws.cell(row=header_row, column=5, value="Funded").font = Font(bold=True)
            ws.cell(row=header_row, column=6, value="Balance").font = Font(bold=True)

            # Determine start date (from vintage_range if available, otherwise use a default)
            if portfolio.selection_criteria.get('vintage_range'):
                start_date, _ = portfolio.selection_criteria['vintage_range']
                if isinstance(start_date, str):
                    start_date = pd.to_datetime(start_date)
            else:
                # Default start date (one month before current date)
                start_date = datetime.now().replace(day=1) - pd.DateOffset(months=1)

            # Generate rows for each month
            for i in range(len(monthly_vintages)):
                # Calculate EOMONTH for the current month
                current_row = header_row + 1 + i
                date_cell = ws.cell(row=current_row, column=4)
                
                # First cell uses direct date formula
                if i == 0:
                    date_formula = f"=EOMONTH(DATE({start_date.year},{start_date.month},1),0)"
                else:
                    # Subsequent cells reference the cell above
                    date_formula = f"=EOMONTH(D{current_row-1},1)"
                
                date_cell.value = date_formula
                date_cell.number_format = "m/d/yyyy"
                
                # Funded formula (sum of allocations by month)
                funded_cell = ws.cell(row=current_row, column=5)
                funded_formula = f"=SUMIFS(Deals!$G:$G,Deals!$C:$C,\">=\"&DATE(YEAR(D{current_row}),MONTH(D{current_row}),1),Deals!$C:$C,\"<=\"&D{current_row})"
                funded_cell.value = funded_formula
                funded_cell.number_format = EXCEL['CURRENCY_FORMAT']
                
                # Balance formula (sum of cashflow amounts as of date)
                balance_cell = ws.cell(row=current_row, column=6)
                balance_formula = f"=-SUMIFS(Cashflows!$E:$E,Cashflows!$A:$A,\"<=\"&$D{current_row})"
                balance_cell.value = balance_formula
                balance_cell.number_format = EXCEL['CURRENCY_FORMAT']
            
        return ws
        
    except Exception as e:
        logger.error(f"Error creating Portfolio Scenario sheet: {str(e)}")
        # Create a basic sheet with error message
        ws = wb.create_sheet("Portfolio Scenario", 0)
        ws.cell(row=1, column=1, value=f"Error creating sheet: {str(e)}")
        return ws


def _create_cashflows_sheet(wb, portfolio):
    """
    Create the Cashflows worksheet with all individual cashflows.
    
    Args:
        wb: Workbook object
        portfolio: Portfolio instance
        
    Returns:
        Worksheet object
    """
    try:
        ws = wb.create_sheet("Cashflows", 1)
        
        if portfolio.cashflows is None or portfolio.cashflows.empty:
            ws.cell(row=1, column=1, value="No cashflow data available")
            return ws
        
        # Set column widths
        _set_column_widths(ws, {
            'A': 15,  # Transaction Date
            'B': 15,  # Funded Date
            'C': 15,  # Funded ID
            'D': 30,  # Transaction Description
            'E': 15,  # Portfolio Amount
            'F': 15,  # Fee Amount
            'G': 15,  # Net Amount
            'H': 10,  # Principal Repaid
        })
        
        # Define columns to include
        columns = [PAYMENT_COLUMNS['DATE'], PAYMENT_COLUMNS['FUNDED DATE'], PAYMENT_COLUMNS['ID'], PAYMENT_COLUMNS['DESCRIPTION'], 
                  'Portfolio Amount', 'Fee Amount', 'Net Amount', 'Principal Repaid']
        
        # Prepare cashflow data
        cf_data = portfolio.cashflows.copy()
        
        # Sort by Transaction Date
        cf_data = cf_data.sort_values(PAYMENT_COLUMNS['DATE'])
        
        # Define column formats - IMPORTANT: Don't convert dates to strings for Excel
        formats = {
            PAYMENT_COLUMNS['DATE']: _format_date_cell,  
            PAYMENT_COLUMNS['FUNDED DATE']: _format_date_cell,  
            'Portfolio Amount': _format_money_cell,
            'Fee Amount': _format_money_cell,
            'Net Amount': _format_money_cell,
            'Principal Repaid': _format_boolean_cell
        }
        
        # Create the table
        _create_table_from_df(
            ws=ws,
            df=cf_data,
            start_row=1,
            columns=columns,
            formats=formats,
            include_totals=False
        )
        
        return ws
        
    except Exception as e:
        logger.error(f"Error creating Cashflows sheet: {str(e)}")
        # Create a basic sheet with error message
        ws = wb.create_sheet("Cashflows", 1)
        ws.cell(row=1, column=1, value=f"Error creating sheet: {str(e)}")
        return ws


def _create_deals_sheet(wb, portfolio):
    """
    Create the Deals worksheet with all individual deals in the portfolio.
    
    Args:
        wb: Workbook object
        portfolio: Portfolio instance
        
    Returns:
        Worksheet object
    """
    try:
        ws = wb.create_sheet("Deals", 2)
        
        if portfolio.deal_allocations is None or portfolio.deal_allocations.empty:
            ws.cell(row=1, column=1, value="No deal data available")
            return ws
        
        # Set column widths
        _set_column_widths(ws, {
            'A': 15,  # Funded ID
            'B': 15,  # Product
            'C': 15,  # Initial Funding Date
            'D': 12,  # Credit Tier
            'E': 15,  # Total Original Balance
            'F': 15,  # Total Original RTR
            'G': 15,  # Commission Cost %
            'H': 15,  # Allocation Amount
            'I': 15,  # Allocation Percentage
            'J': 15,  # Allocated RTR
        })
        
        # Define columns to include
        columns = [
            DEAL_COLUMNS['ID'], DEAL_COLUMNS['PRODUCT'], DEAL_COLUMNS['FUNDING_DATE'], DEAL_COLUMNS['CREDIT_TIER'], 
            DEAL_COLUMNS['BALANCE'], DEAL_COLUMNS['RTR'], DEAL_COLUMNS['COMMISSION'],
            'Allocation Amount', 'Allocation Percentage', 'Allocated RTR'
        ]
        
        # Prepare deal data
        deal_data = portfolio.deal_allocations.copy()
        
        # Sort by Initial Funding Date if available
        if DEAL_COLUMNS['FUNDING_DATE'] in deal_data.columns:
            deal_data = deal_data.sort_values(DEAL_COLUMNS['FUNDING_DATE'])
        
        # Define column formats - IMPORTANT: Don't convert dates to strings for Excel
        formats = {
            DEAL_COLUMNS['FUNDING_DATE']: _format_date_cell,  # Pass datetime objects directly
            DEAL_COLUMNS['BALANCE']: _format_money_cell,
            DEAL_COLUMNS['RTR']: _format_money_cell,
            'Allocation Amount': _format_money_cell,
            'Allocation Percentage': _format_percent_cell,
            DEAL_COLUMNS['COMMISSION']: _format_percent_cell,
            'Allocated RTR': _format_money_cell
        }
        
        # Create the table
        _create_table_from_df(
            ws=ws,
            df=deal_data,
            start_row=1,
            columns=columns,
            formats=formats,
            include_totals=False
        )
        
        return ws
        
    except Exception as e:
        logger.error(f"Error creating Deals sheet: {str(e)}")
        # Create a basic sheet with error message
        ws = wb.create_sheet("Deals", 2)
        ws.cell(row=1, column=1, value=f"Error creating sheet: {str(e)}")
        return ws